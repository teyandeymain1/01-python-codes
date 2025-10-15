#from google.colab import drive
#drive.mount('/content/drive')
import openpyxl as op
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta

#==============================================変数クラス==========================================================================
class varCls(): #変数クラス、すべてクラス変数で定義する。

    #ワークブックのPathの定義
    teacherWbPath = "/Users/teyan/Desktop/ITTO希望日程表/1設定と講師用希望日程表.xlsx"
    studentWbPath = "/Users/teyan/Desktop/ITTO希望日程表/2生徒用希望日程表.xlsx"

    #ワークブックとワークシートの定義
    teacherWb = op.load_workbook(teacherWbPath)
    inTeacherWs = teacherWb["入力画面"]
    teacherVarWs = teacherWb["teacherVar"]
    #設定画面
    variableWs = teacherWb["設定画面"] 

    studentWb = op.load_workbook(studentWbPath)   
    inStudentWs = studentWb["入力画面"]     
    #studentVarWs = studentWb["studentVar"]

    #日付見出しをつくるための変数
    startYear = datetime.today().year
    startMonth = 1 #variableWs.cell(?,?).value
    startDay = 1 #variableWs.cell(?,?).value
    endYear = datetime.today().year
    endMonth = 2 #variableWs.cell(?,?).value
    endDay = 9 #variableWs.cell(?,?).value 
    #datetime.timedelta型をstr型に置き換え->"XX days, 0:00:00"という文字列に変換されるので0~1番目までスライスして(XXのみ取り出して)int型へ
    numOfDays = int(str(datetime(endYear, endMonth, endDay) - datetime(startYear, startMonth, startDay))[:2]) + 1

    #時刻見出しをつくるための変数
    startHour = 13 #variableWs.cell(?,?).value
    startMin = 0 #variableWs.cell(?,?).value
    endHour = 21 #variableWs.cell(?,?).value
    endMin = 0 #variableWs.cell(?,?).value
    lessonTime = 60 #variableWs.cell(?,?).value
    periods = 9 + 1 #variableWs.cell(?,?).value + 1 #授業のコマ数(見出し分一行追加)    

    #表のサイズと形を決める変数
    yMin = 3                #表のスタートの行(固定) 行(row) = y
    xMin = 1                #表のスタートの列(固定) 列(col) = x
    daysInCol = 14 #variableWs.cell(?,?).value #表の列(横)に入れる日数
    xLen = daysInCol + 1    #表の横の長さ(見出し一列分追加)   
    xMax = xLen + xMin - 1  #表の横の最終行(x座標の最大値)
    if numOfDays%daysInCol == 0: numOfRow = numOfDays//daysInCol       #表の束数を決める
    else:                        numOfRow = (numOfDays//daysInCol) + 1 #総日数が表の横の長さで割り切れないときは一行追加
    yLen = numOfRow*periods #表の縦の長さ
    yMax = yLen + yMin - 1  #表の最終列の座標(y座標の最大値)

    #教科を決める
    #subjectList = tuple((variableWs.cell(?, i).value, 0) if variableWs.cell(?+1, i).value == None\
    #                       else (variableWs.cell(?, i).value, variableWs.cell(?+1, i).value)\
    #                           for i in range(variableWs.cell(?,?).value)) 

#==================================================================================================================


#空きコマ検索用辞書を操作するクラス、すべてクラス変数/メソッドで定義する。
class rsvCls():

    #日付および時刻見出しを除く{(y%varCls.periods) != 0 の部分}表(Python上の二次元配列)の座標を記録
    locTuple = tuple((y, x) for y in range(1, varCls.yLen, 1)\
                                for x in range(1, varCls.xLen, 1)\
                                    if (y%varCls.periods) != 0)  #pythonの二次元配列は(0,0)開始なので0行目と0列目が見出し
    rsvTuple = tuple(varCls.teacherVarWs.cell((y+1), (x+1)).value for (y, x) in locTuple) #teacherVarWsの読み取り、エクセルは(1, 1)から始まるため+1する
    rsvDict = dict(zip(locTuple, rsvTuple)) #講師の希望コマを記録する辞書を作成

    # に書き込み
    @classmethod
    def write_rsvDict_into_sheet(cls, outMatrix):
        for (y, x) in cls.rsvDict:
            varCls.teacherVarWs.cell((y+1), (x+1), cls.rsvDict[(y, x)]) #エクセルは(1, 1)から始まるため+1する
        #variableWbに保存
        varCls.teacherWb.save(varCls.teacherWbPath)

    #講師の出勤可能日を見つけてrsvDictに記録
    @classmethod
    def find_available_date(cls, inMatrix):
        for (y, x) in cls.rsvDict:
            if inMatrix[y][x] == 0:
                #講師希望日程表の操作
                bufList = list( "" if cls.rsvDict[(y, x)] == None\
                                    else cls.rsvDict[(y, x)].split("/")) #文字列をlistへ、cls.rsvDictがNoneの時は"N"を要素としてリスト作成 ←ここを文字列のまま処理できないのか？
                bufList.append(varCls.inTeacherWs.cell(1, 2).value)      #講師名を追加、講師名はinTeacherWsのA2セルに記入
                bufString = "/".join(bufList)                            
                cls.rsvDict[(y, x)] = bufString   #"A/B/..."という文字列に変換して代入
        cls.write_rsvDict_into_sheet(cls.rsvDict) #"variableWs"に書き込み
        
    #生徒の希望日をみつけて、rsvDictに記録
    @classmethod
    def reserve_periods(cls, inMatrix, inList):
        for (y, x) in cls.rsvDict:
            if inMatrix[y][x] == 0:
                #講師希望日程表の操作
                bufList = list( "" if cls.rsvDict[(y, x)] == None\
                                    else cls.rsvDict[(y, x)].split("/")) #文字列をlistへ、cls.rsvDictがNoneの時は"N"を要素としてリスト作成 ←ここを文字列のまま処理できないのか？
                if len(bufList) != 0:          
                    bufString = "/".join(bufList[:-1])  #bufList[:-1]で末尾削除 => 講師名の削除、文字列として結合して代入
                    cls.rsvDict[(y, x)] = bufString 
                    #生徒希望日程表の操作
                    inMatrix[y][x] = inList.pop() #教科を代入
                else: #生徒希望日程表の操作
                    inMatrix[y][x] = "講師不足" #エラーコードを代入
        #"variableWs"に書き込み
        cls.write_rsvDict_into_sheet(cls.rsvDict) #"variableWs"に書き込み
        #教科代入後の二次配列を返す
        return inMatrix


#セルを読み書きするクラス、インスタンス変数/メソッドで定義
class read_and_write_class():
    
    #インスタンスの生成&初期化
    def __init__(self, ws) -> None:
        self.ws = ws

    #表を作る
    def make_table(self):
        #シートを初期化
        for _ in range(varCls.yMax+1): #既存の表を1行ずつ削除
            self.ws.delete_rows(1)
        #線の仕様
        thinLine = Side(style='thin', color='000000')
        thickLine = Side(style='thick', color='000000')
        #罫線を引く
        for y in range(varCls.yMin, (varCls.yMax+1), 1):     #
            for x in range(varCls.xMin, (varCls.xMax+1), 1): #
                if (y-varCls.yMin)%varCls.periods == 0 and (x-varCls.xMin) == 0: #"時\日"のセルは四方を最太線
                    self.ws.cell(y, x).border = Border(top=thickLine, bottom=thickLine, left=thickLine, right=thickLine)             
                elif (y-varCls.yMin)%varCls.periods == 0:                        #日付の行はセルの上下を最太線
                    self.ws.cell(y, x).border = Border(top=thickLine, bottom=thickLine, left=thinLine, right=thinLine)                
                elif (x-varCls.xMin) == 0:                                       #時刻の列はセルの左右を最太線 
                    self.ws.cell(y, x).border = Border(top=thinLine, bottom=thinLine, left=thickLine, right=thickLine)                    
                else:                                                            #その他のセルは四方を最細線
                    self.ws.cell(y, x).border = Border(top=thinLine, bottom=thinLine, left=thinLine, right=thinLine)     
        #端の罫線の処理
        for y in range(varCls.yMin, (varCls.yMax+1), 1): #右端の境界線
            self.ws.cell(y, (varCls.xMax+1)).border = Border(left=thickLine)
        for x in range(varCls.xMin, (varCls.xMax+1), 1): #下端の境界線
            self.ws.cell((varCls.yMax+1), x).border = Border(top=thickLine)
        #土日に色付け
        for y in range((varCls.yMin), (varCls.yMax+1), varCls.periods): #
            for x in range((varCls.xMin+1), (varCls.xMax+1), 1):        #ただし、(x-varCls.xMin)=0の時を除くために(varCls.xMin+1)から開始
                if (x-varCls.xMin)%6 == 0 or (x-varCls.xMin)%7 == 0:
                    self.ws.cell(y, x).fill = PatternFill(patternType='solid', fgColor='00C0C0C0') #色の詳細は https://openpyxl.readthedocs.io/en/stable/styles.html#indexed-colours

    #セルを読み込む
    def read_sheet_make_matrix(self): # iter_rows() で行ごと取得
        return list(list(value for value in rowList)\
                        for rowList in self.ws.iter_rows(min_row=varCls.yMin, min_col=varCls.xMin, max_row=varCls.yMax, max_col=varCls.xMax, values_only=True))

    #セルに書き込む   
    def write_matrix_into_sheet(self, inMatrix): #この関数はクラス外部から直接的に二次元配列を引数として受け取る。
        for y, rowList in enumerate(inMatrix, varCls.yMin):
            for x, value in enumerate(rowList, varCls.xMin):
                self.ws.cell(y, x, value)


#初期状態の表作成クラス、インスタンス変数/メソッドで定義
class make_initial_table_class(): 
    
    #インスタンスの生成&初期化
    def __init__(self, ws) -> None:
        self.ws = ws

    def make_initial_matrix(self): #
        #日付の自動生成
        dateList = [datetime(varCls.startYear, varCls.startMonth, varCls.startDay) + timedelta(days=i) for i in range(varCls.numOfDays)]
        dateList = [day.strftime("%m/%d") for day in dateList]
        dateList = [["時\日"] + dateList[d:(d+varCls.daysInCol)] for d in range(0, varCls.numOfDays, (varCls.daysInCol))]
        #時刻の自動生成
        timeList = [datetime(1, 1, 1, varCls.startHour, varCls.startMin) + timedelta(minutes=(varCls.lessonTime*t)) for t in range((varCls.periods-1))]
        timeList = (["時\日"] + [time.strftime("%H:%M") for time in timeList])*varCls.numOfRow                                #periods-1は見出し分の一行を除くため
        #表と対応する二次元配列
        matrixToSheet = [[None]*(varCls.xLen) for _ in [None]*varCls.yLen]
        #日付見出し代入
        for y in range(varCls.numOfRow):
            matrixToSheet[(y*varCls.periods)] = dateList[y] #(y*varCls.periods)ずつ行を飛ばしながら日付見出しを代入
        #時刻見出し代入
        for y, rowList in enumerate(matrixToSheet):
            rowList[0] = timeList[y]    
        return matrixToSheet

    #講師用希望日程表の初期化
    def make_initial_teacher_table(self):  
        #初期状態の表を定義
        matrixToSheet = self.make_initial_matrix()
        #貼り付けクラス&関数の呼出
        rwCls = read_and_write_class(self.ws)        #インスタンス化
        rwCls.make_table()                           #罫線を引く 
        rwCls.write_matrix_into_sheet(matrixToSheet) #Python上の二次元配列をシートに書き込み
        #上書き保存
        varCls.teacherWb.save(varCls.teacherWbPath)
        #講師控え用のシートを作成
    
    #生徒用希望日程表の初期化
    def make_initial_student_table(self):
        #初期状態の表を定義
        matrixToSheet = self.make_initial_matrix()
        #満席のコマを"X"で埋める
        for (y, x) in rsvCls.rsvDict:
            if rsvCls.rsvDict[(y, x)] == None:
                matrixToSheet[y][x] = "X"
            else:
                matrixToSheet[y][x] = rsvCls.rsvDict[(y, x)]
        #貼り付けクラス&関数の呼出
        rwCls = read_and_write_class(self.ws)      
        rwCls.make_table()                                  
        rwCls.write_matrix_into_sheet(matrixToSheet)
        #上書き保存
        varCls.studentWb.save(varCls.studentWbPath)  

#class 

    #インスタンスの生成&初期化
    #def __init__(self, ws) -> None:
        #self.ws = ws

    #教科格納配列の作成&python二次元配列に代入
    #def (self, inMatrix): 
        #for subject, numOfSub in varCls.subjectList:
            #inList.extend([subject]*numOfSub)
            #inListをランダムに入れ替える
        #rsvDictの人数減らす&シートに貼り付け
        #matrixToSheet = varCls.reserve_periods(cls, inMatrix, inList)
        #貼り付けクラス&関数の呼出
        #rwCls = read_and_write_class(self.ws)
        #rwCls.write_matrix_into_sheet(matrixToSheet)
        #上書き保存
        #varCls.studentWb.save(varCls.studentWbPath)  
